#!/usr/bin/env python3
# -*-encoding: utf-8-*-

from tkinter import *
from tkinter.messagebox import showwarning
import datetime
from openpyxl import load_workbook


def get_base(filename='rates.xlsx', sheetname='List1'):
    """ Завантажує інформацію про курс валют з таблиці exel

    :param filename: назва таблиці
    :param sheetname: назва робочого аркуша
    :return: словник {(валюта1, валюта2): курс}
    """
    wb = load_workbook(filename)
    ws = wb[sheetname]

    res = {}
    n = ws.max_row     # номер останнього рядка
    for i in range(2, n+1):
        tmp1 = ws['A{}'.format(i)].value    # проходимо по всіх рядках і зчитуємо 3 поля
        tmp2 = ws['B{}'.format(i)].value
        rate = ws['C{}'.format(i)].value
        if not all([tmp1, tmp2, rate]):     # якщо хоч одне поле пусти - пропускаємо рядок
            continue
        res[(tmp1, tmp2)] = rate            # добавляємо в словник знайдену пару
        if (tmp2, tmp1) not in res:         # якщо в словнику відсутня протилежна пара
            res[(tmp2, tmp1)] = 1/float(rate)    # то добавляємо в словник її, а курс робимо оберненим
    return res


class GUI:
    """ Графічний інтерфейс.
        Головне вікно містить 2 списки - доступні валюти.
        Вибір валюти реалізується подвійним натисненням.
    """
    _pattern = "Результат: {}"    # шаблон виведення результату

    def __init__(self, master):
        """ Ініціалізація

        :param master: вікно, до якого прив'язаний графічний інтерфейс
        """
        self.top = master
        self.filename = 'data.xlsx'     # ексель таблиця
        self.sheet = 'List1'            # аркуш

        self._exchange = None        # словник курсу валют
        self._update_base()          # заванажити словник
        self._make_widgets()         # зробити віджети
        self._from = None
        self._to = None

    def _update_base(self):
        """ Завантажує курси валют з ексель таблиці.
        """
        self._exchange = get_base(self.filename, self.sheet)

    def _make_widgets(self):
        """ Створення віджетів
        """
        # написи у верхній частині вікна
        self.frame_labels = Frame(self.top)
        Label(self.frame_labels, text='    FROM', font=('Helvetica', '12', 'bold')).pack(side=LEFT)
        Label(self.frame_labels, text='TO    ', font=('Helvetica', '12', 'bold')).pack(side=RIGHT)
        self.frame_labels.pack(side=TOP, fill=X)

        self._make_lists()    # списки валют

        # вибрані валюти (написи в нижній частині вікна)
        self.frame_chosen = Frame(self.top)
        Label(self.frame_chosen, text='From', font=('arial', '16', 'bold')).pack(side=LEFT)
        self.output1 = Label(self.frame_chosen, font=('arial', '16'))
        self.output1.pack(side=LEFT)
        Label(self.frame_chosen, text='TO', font=('arial', '16', 'bold')).pack(side=LEFT)
        self.output2 = Label(self.frame_chosen, font=('arial', '16'))
        self.output2.pack(side=LEFT)

        self.frame_chosen.pack(side=TOP, fill=X)

        # рядок введення суми
        Label(self.top, text='Сума:', font=('arial', '16')).pack(side=LEFT)
        self.input = Entry(self.top, font=('arial', '16'))
        self.input.pack(side=LEFT)

        # рядок виведення результату
        self._res = Label(self.top, text=GUI._pattern.format(''), font=('arial', '16'))
        self._res.pack(side=LEFT)
        # кнопка запуску
        Button(self.top, text='Convert', font=('Helvetica', '16'), command=self._convert_handler).pack(side=RIGHT)

    def _make_lists(self):
        """ Створення 2-х списків у рамці
        """
        self._lists = Frame(self.top)
        self._from_frame = Frame(self._lists)
        self._from_scroll = Scrollbar(self._from_frame)
        self._from_scroll.pack(side=RIGHT, fill=Y)
        self._from_list = Listbox(self._from_frame, height=15,
                                  width=50, yscrollcommand=self._from_scroll.set)
        self._from_scroll.config(command=self._from_list.yview)
        self._from_list.pack(side=RIGHT, fill=BOTH, expand=YES)
        self._from_frame.pack(side=LEFT, fill=Y, expand=YES)

        self._from_list.bind('<Double-1>', self._left_handler)

        self._to_frame = Frame(self._lists)
        self._to_scroll = Scrollbar(self._to_frame)
        self._to_scroll.pack(side=RIGHT, fill=Y)
        self._to_list = Listbox(self._to_frame, height=15,
                                width=50, yscrollcommand=self._to_scroll.set)
        self._to_scroll.config(command=self._to_list.yview)
        self._to_list.pack(side=RIGHT, fill=BOTH, expand=YES)
        self._to_frame.pack(side=LEFT, fill=Y, expand=YES)

        self._to_list.bind('<Double-1>', self._right_handler)

        self._fill_lists()
        self._lists.pack(side=TOP, fill=X, expand=YES)

    def _fill_lists(self):
        """ Наповнення списків інформацією
        """
        for cur_from, cur_to in self._exchange:     # проходимо по ключах
            self._from_list.insert(END, cur_from)   # перший елемент ключа - валюта з якої відбувається конвертування
            self._to_list.insert(END, cur_to)       # другий елемент - в яку
        self.top.update()

    def _left_handler(self, ev=None):
        """ Обробити подвійне натиснення лівої клавіші миші для лівого списку"""
        try:
            # отримати вибраний елемент списку
            check = self._from_list.get(self._from_list.curselection())
            self.output1.config(text=check)    # записуємо вибране значення в напис у нижній частині вікна
            self._from = check                 # запам'ятовуємо значення

        except TclError:
            # пропустити помилку curselection, якщо під час
            # подвійного натиснення лівої клавіші миші список порожній
            pass
        except Exception as e:
            # якщо інша помилка, то видати повідомлення
            showwarning('Помилка', e)
            raise e                   # і виключення для логування у файл

    def _right_handler(self, ev=None):
        """ Обробити подвійне натиснення лівої клавіші миші для правого списку.
            Все аналогічно.
        """
        try:
            # отримати вибраний елемент списку
            check = self._to_list.get(self._to_list.curselection())
            self.output2.config(text=check)    # записати значення в низ вікна
            self._to = check                   # запам'ятати його в змінній
        except TclError:
            # пропустити помилку curselection, якщо під час
            # подвійного натиснення лівої клавіші миші список порожній
            pass
        except Exception as e:
            # якщо інша помилка, то видати повідомлення
            showwarning('Помилка', e)
            raise e                     # і виключення для логування у файл

    def _convert_handler(self, ev=None):
        """ Обробити натиснення лівої клавіші миші на кнопку.
        """
        try:
            # зчитуємо інформацію з поля для введення
            summ = self.input.get()

            # якщо поле не заповнене, або не вибрані курси валют, то нічого не треба робити
            if not summ or self._from is None or self._to is None:
                return

            # вибираємо з словника курс для вибраної пари
            coeff = self._exchange[(self._from, self._to)]
            res = float(summ) * float(coeff)   # обчислення
            self._res.config(text=GUI._pattern.format(res))  # виведення
            self.top.update()

        except TclError:
            # пропустити помилку curselection, якщо під час
            # подвійного натиснення лівої клавіші миші список порожній
            pass
        except KeyError:
            showwarning('Помилка',
                        'в базі даних нема інформації про курс обміну {} на {}'.format(self._from, self._to))
        except Exception as e:
            # якщо інша помилка, то видати повідомлення
            showwarning('Помилка', e)
            raise e                         # і виключення для логування у файл


if __name__ == '__main__':
    import sys

    # перевизначаємо стандартний потік виключень, щоб помилки виводились у файл
    sys.stderr = open('gui.log', 'a', encoding='utf-8')
    sys.stderr.write('\n\n-----------{}----------\n'.format(datetime.datetime.now()))

    top = Tk()
    test = GUI(top)
    mainloop()

    sys.stderr.close()   # перед виходом з програми обов'язково закрити файл
