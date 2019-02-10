#!/usr/bin/env python3
# -*-encoding: utf-8-*-

import openpyxl
import datetime
from collections import namedtuple


Flight = namedtuple('Flight', 'id depart arrive flight_class cost')


def get_all_airports(filename: str, sheetname='Airports'):
    """ Повертає всі аеропорти з exel таблиці

    :param filename: назва таблиці
    :param sheetname: назва аркуша
    :return: словник {назва аеропорта: id}
    """
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheetname]
    res = {}
    n = ws.max_row
    for i in range(2, n+1):
        res[ws[f'B{i}'].value] = ws[f'A{i}'].value
    return res


def get_all_flights(filename: str, depart_id, arrive_id, date: tuple, sheetname='Flights'):
    """ Повертає всі рейси з одного аеропорту
        в інший на задану дату

    :param filename: назва файлу
    :param depart_id: id аеропорту-відправлення
    :param arrive_id: id аеропорту-прибуття
    :param date: дата tuple(рік, місяць, день)
    :param sheetname: назва аркуша
    :return: список іменованих кортежів
    """
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheetname]
    n = ws.max_row
    res = []
    date_day = datetime.datetime(*date).weekday()   # день тижня, на який припадає задана дата
    for i in range(2, n+1):
        from_ = ws[f'A{i}'].value
        to_ = ws[f'B{i}'].value                        # перевіряємо задані умови
        days = str(ws[f'D{i}'].value)                  # і додаємо до списку ті рейси, які їх задовольняють
        if from_ in depart_id and to_ in arrive_id and days[date_day] != '0':
            res.append(Flight(ws[f'C{i}'].value, ws[f'E{i}'].value,
                              ws[f'F{i}'].value, ws[f'G{i}'].value,
                              ws[f'H{i}'].value))
    return res
