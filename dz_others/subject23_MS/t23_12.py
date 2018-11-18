#!/usr/bin/env python3
# -*-encoding: utf-8-*-

# by David Zashkolny
# 2 course, comp math
# Taras Shevchenko National University of Kyiv
# email: davendiy@gmail.com

import openpyxl
import datetime


def get_id_city(filename: str, city: str, sheetname='Airports'):
    """ Повертає множину id всіх аеропортів деякого міста з
        exel-таблиці, яка має вид даної у завданні

    :param filename: назва таблиці
    :param city: назва міста
    :param sheetname: назва робочого аркуша
    :return: множина всіх id
    """
    wb = openpyxl.load_workbook(filename)  # відкриваємо таблицю
    ws = wb[sheetname]                     # відкриваємо робочий аркуш
    city = city.lower()                    # переводимо назву до нижнього регістру

    res = set()        # результуюча множина
    n = ws.max_row     # к-ть рядків у таблиці
    for i in range(1, n):
        if city in ws[f'C{i+1}'].value.lower():  # проходимо по всіх рядках і перевіряємо назву міста
            res.add(ws[f'A{i+1}'].value)         # якщо знайшли, то додаємо до множини
    return res


def find_cheapest(filename: str, from_ids, to_ids, sheetname='Flights'):
    """ Знаходження найдешевшого рейсу в таблиці exel, яка має вид даної у
        завданні

    :param filename: назва таблиці
    :param from_ids: id всіх аеропортів міста відравлення
    :param to_ids: id всіх аеропортів міста прибуття
    :param sheetname: назва робочого аркуша
    :return: кортеж (id рейса, дні відправлення, час відправлення, час прибуття, клас, вартість)
    """
    wb = openpyxl.load_workbook(filename)  # відкриваємо таблицю
    ws = wb[sheetname]                     # відкриваємо робочий аркуш

    n = ws.max_row  # к-ть рядків
    cheapest = 0    # найдешевший рейс
    res_i = 0       # номер найдешевшого рейсу
    flag = True     # флаг того, що знайшли елемент перший раз

    for i in range(1, n):
        # перевіряємо, чи id аеропортів містяться у вхідних структурах
        if ws[f'A{i+1}'].value not in from_ids or ws[f'B{i+1}'].value not in to_ids:
            continue
        tmp = float(ws[f'H{i+1}'].value)
        if flag or tmp < cheapest:  # якщо знайшли перший раз, або знайшли менший,
            cheapest = tmp          # то переприсваюємо значення найменшого
            flag = False
            res_i = i + 1

    if cheapest:  # якщо знайшли найдешевший рейс (його може не бути зовсім),
        # то зчитуємо всю необхідну інформацію
        res = [ws[f'{symbol}{res_i}'].value for symbol in 'ABCDEFGH']
    else:
        res = None
    return res


def find_fastest(filename, from_ids, to_ids, sheetname='Flights'):
    """ Знаходження найшвидшого рейсу в таблиці exel, яка має вид даної у
            завданні

        :param filename: назва таблиці
        :param from_ids: id всіх аеропортів міста відравлення
        :param to_ids: id всіх аеропортів міста прибуття
        :param sheetname: назва робочого аркуша
        :return: кортеж (id рейса, дні відправлення, час відправлення, час прибуття, клас, вартість)
        """
    wb = openpyxl.load_workbook(filename)  # відкриваємо таблицю
    ws = wb[sheetname]  # відкриваємо робочий аркуш

    n = ws.max_row   # к-ть рядків
    fastest = 0      # найдешевший рейс
    res_i = 0        # номер найдешевшого рейсу
    flag = True      # флаг того, що знайшли елемент перший раз

    for i in range(1, n):     # проходимо по всіх рейсах
        # перевіряємо, чи рейс з'єднує необхідні нам міста
        if ws[f'A{i+1}'].value not in from_ids or ws[f'B{i+1}'].value not in to_ids:
            continue

        depart = str(ws[f'E{i+1}'].value).split(':')   # зчитуємо час відправлення і прибуття
        arrive = str(ws[f'F{i+1}'].value).split(':')

        # створюємо з них об'єкти datetime, з стандартною датою (їх можна віднімати)
        depart = datetime.datetime(1999, 1, 1, int(depart[0]), int(depart[1]))
        arrive = datetime.datetime(1999, 1, 1, int(arrive[0]), int(arrive[1]))
        if arrive < depart:     # якщо час прибуття раніше, ніж час відправлення, то прибуття буде
            arrive = datetime.datetime(1999, 1, 2, arrive.hour, arrive.minute)    # наступного дня

        flight_time = arrive - depart      # знаходимо тривалість рейсу
        if flag:
            fastest = flight_time
            res_i = i + 1
            flag = False

        elif flight_time < fastest:  # перевіряємо чи він менший за мінімальний
            fastest = flight_time
            res_i = i + 1

    if fastest:     # якщо знайшли, то аналогічно зчитуємо всю необхідну інформацію
        res = [ws[f'{symbol}{res_i}'].value for symbol in 'ABCDEFGH']
    else:
        res = None
    return res


if __name__ == '__main__':
    tablename = 't23_12_input.xlsx'     # будемо шукати рейси з Києва до Парижа
    depart_city = 'Київ'
    arrive_city = 'Париж'

    # спочатку знаходимо ід всіх аеропортів
    depart_ids = get_id_city(tablename, depart_city, sheetname='Аеропорти')
    arrive_ids = get_id_city(tablename, arrive_city, sheetname='Аеропорти')

    # потім знаходимо найдешевший і найшвидший
    res1 = find_cheapest(tablename, depart_ids, arrive_ids, sheetname='Рейси')
    res2 = find_fastest(tablename, depart_ids, arrive_ids, sheetname='Рейси')

    # шаблон виведення
    res_str = '''
              From ID   - {}
              To ID     - {}
              Flight ID - {}
              Days      - {}
              Depart    - {}
              Arrive    - {}
              Class     - {}
              Cost      - {}
              '''

    print(f"Найдешевший рейс з {depart_city} до {arrive_city}:")
    tmp_str = res_str.format(*tuple(res1)) if res1 is not None else ''
    print(tmp_str)

    print(f"\nНайшвидший рейс з {depart_city} до {arrive_city}:")
    tmp_str = res_str.format(*tuple(res2)) if res2 is not None else ''
    print(tmp_str)
