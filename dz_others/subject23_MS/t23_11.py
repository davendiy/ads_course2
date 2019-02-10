#!/usr/bin/env python3
# -*-encoding: utf-8-*-

import openpyxl


def get_common(projects1: list, projects2: list):
    """ функція, що повертає спільні елементи для двох списків

    :param projects1: список
    :param projects2: список
    :return: список
    """
    res = []
    for el in projects1:
        if el in projects2:
            res.append(el)
            projects2.remove(el)
    return res


def get_dict(filename, sheet_name=None):
    """ Функція загружає в словник всю інформацію з MS таблиці
        В таблиці має бути 2 поля: Project, Person
    :param filename: назва таблиці
    :param sheet_name: назва робочого аркуша
    :return: словник {ім'я: список проектів}
    """
    res = {}
    wb = openpyxl.load_workbook(filename)       # завантажуємо таблицю
    if sheet_name is not None:                  # відкриваємо робочий аркуш, якщо він вказаний
        ws = wb[sheet_name]
    else:
        ws = wb.active                          # інакше - відкриваємо активний
    n = ws.max_row                              # к-ть рядків з інформацією
    for i in range(n):
        project = ws['A{}'.format(i+1)].value   # проходимось по рядкам і зчитуємо відповідно назву проекта
        person = ws['B{}'.format(i+1)].value    # і особи, яка ним займалась

        if person in res:                       # додаємо нашу інформацію в словник, в якому
            res[person].append(project)         # ключами є ім'я особи, а значення - список проектів
        else:
            res[person] = [project]
    return res


def create_table(filename: str, dictionary: dict):
    """ Функція створює нову таблицю і заповнює її елементами з словника

    :param filename: назва файлу, куди буде збережена таблиця
    :param dictionary: словник
    """
    wb = openpyxl.Workbook()         # відкриваємо таблицю
    ws = wb.active                   # відкриваємо активний робочий аркуш
    used = []            # список використаних пар імен (щоб не повторюватись)
    i = 1
    for person1 in dictionary.keys():       # проходимо подвійним циклом по ключам
        for person2 in dictionary.keys():
            if person1 == person2 or {person1, person2} in used:   # перевіряємо, чи не використовували ми
                continue                                           # таку пару

            used.append({person1, person2})                # додаємо пару до використаних

            # знаходимо спільні проекти
            common_projects = get_common(dictionary[person1][:], dictionary[person2][:])

            # записуємо в таблицю
            ws['A{}'.format(i)].value = person1
            ws['B{}'.format(i)].value = person2
            ws['C{}'.format(i)].value = ', '.join(common_projects)
            ws['D{}'.format(i)].value = len(common_projects)
            i += 1
    wb.save(filename)        # зберігаємо


if __name__ == '__main__':
    test_dict = get_dict('t23_11_input.xlsx', 'Лист1')
    create_table('t23_11_output.xlsx', test_dict)
