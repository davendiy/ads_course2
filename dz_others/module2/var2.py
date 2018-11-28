#!/usr/bin/env python3
# -*-encoding: utf-8-*-

# created: 28.11.18
# by David Zashkolny
# 2 course, comp math
# Taras Shevchenko National University of Kyiv
# email: davendiy@gmail.com


import openpyxl


def is_identical(filename1, filename2):
    ws1 = openpyxl.load_workbook(filename1)
    ws2 = openpyxl.load_workbook(filename2)

    if len(ws1.get_sheet_names()) != len(ws2.get_sheet_names()):
        return False

    for name1, name2 in zip(ws1.get_sheet_names(), ws2.get_sheet_names()):
        if name1 != name2:
            return False

    for wb_name in ws1.get_sheet_names():
        wb1 = ws1[wb_name]
        wb2 = ws2[wb_name]

        n1 = wb1.max_row
        n2 = wb2.max_row
        m1 = wb1.max_column
        m2 = wb2.max_column

        if n1 != n2 or m1 != m2 :
            return False

        for i in range(1, n1):
            for j in range(1, m1):
                if wb1.cell(row=i, column=j).value() != wb2.cell(row=i, column=j):
                    return False

    return True


if __name__ == '__main__':
    print(is_identical('input1.xlsx', 'input2.xlsx'))
