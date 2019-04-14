#!/usr/bin/env python3
# -*-encoding: utf-8-*-

import sqlite3
import openpyxl
from .constants import *


def name_dict(dicts_list) -> dict:
    """ Функція, яка з списку словників (дані з бази даних) формує
    словник {Name: id}

    :param dicts_list: [dict('field1': 'val1' ... ), dict('field1': 'val1' ... )]
    :return: dict(Name1: id1, Name2: id2)
    """
    res = {}
    for el in dicts_list:
        res[el['Name']] = el['Id']
    return res


def id_dict(dicts_list):
    """ Функція, яка з списку словників (дані з бази даних) формує
    словник {id: Name}

    :param dicts_list: [dict('field1': 'val1' ... ), dict('field1': 'val1' ... )]
    :return: dict(id1: Name1, id2: Name2)
    """
    res = {}
    for el in dicts_list:
        res[el['Id']] = el['Name']
    return res


def create_xlsx(outfile, data):
    """ Запис результату в exel таблицю

    :param outfile: назва таблиці
    :param data: список з словників [dict('field1': 'val1' ... ), dict('field1': 'val1' ... )] -
                 інформація з бази даних.
    """
    wb = openpyxl.Workbook()      # записуємо дані так у форматі t23_21 (приклад з лекцій)
    ws = wb.active
    names = []
    for i, el in enumerate(data[0], start=1):
        c = ws.cell(row=1, column=i)
        c.value = el
        names.append(el)

    for i, row in enumerate(data, start=2):
        for j, column in enumerate(names, start=1):
            c = ws.cell(row=i, column=j)
            c.value = row[column]
    wb.save(outfile)


class ConnectorDB:
    """Клас з'єднання з базою даних.

    self.urn - розташування БД
    self.conn - об'єкт зв'язку з базою даних
    """

    def __init__(self, urn):
        self.urn = urn
        self.conn = None

    def get_cursor(self):
        """Повертає об'єкт курсор."""
        self.conn = sqlite3.connect(self.urn)  # зв'язатись з БД
        return self.conn.cursor()

    def close(self):
        """Завершує з'єднання з БД."""
        if self.conn:
            self.conn.commit()
            self.conn.close()
        self.conn = None

    def get_fields(self, tablename: str):
        """ Отримати назви полів з таблиці

        :param tablename: назва таблиці
        :return: ['field1', 'field2', ...]
        """
        curs = self.get_cursor()
        curs.execute("SELECT * from {}".format(tablename))
        res = [desc[0] for desc in curs.description]
        self.close()
        return res

    def get_data_dicts(self, query, *parameters, n=DEFAULT_N) -> list:
        """Повертає список словників з даними.

        Як відповідь на запит query з параметрами param.
        :param query: запит SqLite
        :param parameters: змінні параметри запиту, якщо є
        :param n: к-ть елеменів, які повернуться
        :return: [{name_field1: value1, name_field2: value2 ...}, ...]
        """
        curs = self.get_cursor()
        curs.execute(query, parameters)
        # взято з Mark Lutz - Programming Python.
        # отримати назви полів
        colnames = [desc[0] for desc in curs.description]
        # створити список словників
        rowdicts = [dict(zip(colnames, row)) for row in curs.fetchmany(n)]
        self.close()
        return rowdicts

    def get_one_result(self, query, *parameters):
        """ Повернути однин результат запиту

        :param query: запит SqLite
        :param parameters: змінні параметри запиту, якщо є
        :return: те саме, що і curs.fetchone()
        """
        curs = self.get_cursor()
        curs.execute(query, parameters)
        result = curs.fetchone()
        self.close()
        return result


# noinspection SqlResolve
class Storage:
    """ Клас, який забеспечує інтерфейс зв'язку з базою даних.
    """

    def __init__(self, connector: ConnectorDB):
        self.db = connector

    def get_items(self, item_type=KEY_WORD, category='', n=DEFAULT_N):
        """ Повернути елементи з бази даних певного типу і категорії

        :param item_type: тип елемента (KEY_WORD, SITE, LINK)
        :param category: id категорії
        :param n: к-ть елементів, які повернуться
        :return: [dict{field: value}, dict{field:value} ... ]
        """
        assert item_type in [KEY_WORD, SITE, LINK], 'bad item_type'
        # якщо категорія не вказана, то просто повертає всі елементи
        if category:
            query = 'SELECT * FROM {} WHERE Category_id=?'.format(item_type)
            res = self.db.get_data_dicts(query, category, n=n)
        else:
            query = 'SELECT * FROM {}'.format(item_type)
            res = self.db.get_data_dicts(query, n=n)
        return res

    def change_item_id(self, item_id, item_type=KEY_WORD, **params):
        """ Змінити елемент у базі даних по id

        :param item_id: ідентифікатор елемента
        :param item_type: тип елемента (KEY_WORD, SITE, LINK)
        :param params: набір параметрів, які необхідно вставити
        """
        assert item_type in [KEY_WORD, SITE, LINK], 'bad item_type'
        tmp_params = list(params.items())
        query = 'UPDATE {} SET '.format(item_type) \
                + ', '.join(['{}=?'.format(key[0]) for key in tmp_params]) \
                + ' WHERE Id=?'

        parameters = [el[1] for el in tmp_params]
        parameters.append(item_id)
        curs = self.db.get_cursor()
        curs.execute(query, parameters)
        self.db.close()

    def change_link(self, link, **params):
        tmp_params = list(params.items())

        query = 'UPDATE Links SET ' \
                + ', '.join(['{}=?'.format(key[0]) for key in tmp_params]) \
                + ' WHERE Link=?'

        parameters = [el[1] for el in tmp_params]
        parameters.append(link)
        curs = self.db.get_cursor()
        curs.execute(query, parameters)
        self.db.close()

    def del_item(self, item_id, item_type=KEY_WORD):
        """ Видалити елемент з бази даних

        :param item_id: ідентифікатор елемента
        :param item_type: тип елемента (KEY_WORD, SITE, LINK)
        """
        assert item_type in [KEY_WORD, SITE, LINK], 'bad item_type'

        query = 'DELETE FROM {} WHERE Id=?'.format(item_type)
        curs = self.db.get_cursor()
        curs.execute(query, (item_id,))
        self.db.close()

    def add_item(self, item_type, **params):
        """ Додати елемент до бази даних

        :param item_type: тип елемента (KEY_WORD, SITE, LINK)
        :param params: набір параметрів, які необхідно вставити
        """

        assert item_type in [KEY_WORD, SITE, LINK], 'bad item_type'

        # словник -> список кортежів для того, щоб не змінювався порядок елементів
        tmp_params = list(params.items())

        # формуємо запит виду INSERT INTO table (param1, param2) values (?, ?)
        query = 'INSERT INTO {}'.format(item_type) \
                + '(' + ', '.join([el[0] for el in tmp_params]) + ')' \
                + ' values (' + ('?, '*len(tmp_params)).strip(', ') + ")"
        parameters = [el[1] for el in tmp_params]   # значення параметрів
        # print(query)
        curs = self.db.get_cursor()
        curs.execute(query, parameters)
        self.db.close()

    def add_category(self, name):
        """ Додати категорію

        :param name: назва нової категорії
        """
        curs = self.db.get_cursor()
        curs.execute('INSERT INTO Categories (Name) values (?)', (name,))
        self.db.close()

    def del_category(self, name):
        """ Видалити категорію за назвою і всі елементи цієї категорії

        :param name: назва категорії
        """
        _id = self.get_category_id(name)
        if _id:          # якщо знайшли категорію з таким ім'ям, то видаляємо
            curs = self.db.get_cursor()

            # необхідно видалити елементи кожного типу
            for item_type in [KEY_WORD, SITE, LINK]:
                query = 'DELETE FROM {} WHERE Category_id=?'.format(item_type)
                curs.execute(query, (_id,))

            # видаляємо власне категорію
            curs.execute('DELETE FROM Categories WHERE Id=?', (_id,))
            self.db.close()

    def get_category_id(self, name):
        """ Повертає ідентифікатор категорії за назвою

        :param name: рядок - назва
        :return: рядок - ід
        """
        _id = self.db.get_one_result('SELECT Id FROM Categories WHERE Name=?', name)
        return '' if not _id else _id[0]

    def get_categories(self):
        return self.db.get_data_dicts('SELECT * FROM Categories')


data_conn = ConnectorDB(DEFAULT_DATABASE)
database = Storage(data_conn)
