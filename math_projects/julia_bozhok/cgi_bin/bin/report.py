#!/usr/bin/env python3
# -*-encoding: utf-8-*-

""" Модуль з функціями створення звіту і накладної
за шаблоном у MS Word.
"""

from .merge import *
from .constants import *
import openpyxl
import os


def create_xlsx(outfile, data):
    """ Створення тимчасової exel таблиці

    :param outfile: назва таблиці
    :param data: список з словників [dict('field1': 'val1' ... ), dict('field1': 'val1' ... )] -
                 інформація з бази даних.
    """
    wb = openpyxl.Workbook()      # записуємо дані так у форматі t23_21 (приклад з лекцій)
    ws = wb.active
    names = []
    for i, el in enumerate(data[0], start=1):
        c = ws.cell(row=1, column=i)
        c.value = el
        names.append(el)

    for i, row in enumerate(data, start=2):
        for j, column in enumerate(names, start=1):
            c = ws.cell(row=i, column=j)
            c.value = row[column]
    wb.save(outfile)


def create_report(outfile, template, data):
    """ Створення звіту, або накладної за шаблоном у MS Word.

    Задача аналогічна прикладу t23_21 з лекцій, де необхідно
    було злити декілька файлів в один за шаблоном. Тому просто створюємо
    ексель таблицю з даними і використовуємо функціонал прикладу з
    лекцій.

    :param outfile: шлях до вихідного файлу
    :param template: шлях до шаблону
    :param data: [dict('field1': 'val1' ... ), dict('field1': 'val1' ... )]
    """
    params = dict([(name, TMP_FILE_NAME) for name in data[0]])
    lead_param = 'id'
    create_xlsx(TMP_FILE_NAME, data)
    m = Merger(template, params, lead_param, outfile)
    m.merge()
    path = os.path.curdir
    os.remove(os.path.join(path, TMP_FILE_NAME))
