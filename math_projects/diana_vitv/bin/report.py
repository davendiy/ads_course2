#!/usr/bin/env python3
# -*-encoding: utf-8-*-

from .merge import *
import openpyxl
import os

TMP_FILE_NAME = '___tmp.xlsx'


def create_xlsx(outfile, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    names = []
    for i, el in enumerate(data[0], start=1):
        c = ws.cell(row=1, column=i)
        c.value = el
        names.append(el)

    for i, row in enumerate(data, start=2):
        for j, column in enumerate(names, start=1):
            c = ws.cell(row=i, column=j)
            c.value = row[column]
    wb.save(outfile)


def create_report(outfile, template, data):
    params = dict([(name, TMP_FILE_NAME) for name in data[0]])
    lead_param = 'id'
    create_xlsx(TMP_FILE_NAME, data)
    m = Merger(template, params, lead_param, outfile)
    m.merge()
    path = os.path.curdir
    os.remove(os.path.join(path, TMP_FILE_NAME))


def create_invoice():
    pass